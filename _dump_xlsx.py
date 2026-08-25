import openpyxl

path = r'C:\Users\Administrator\Desktop\角色权限矩阵与审批流程配置表.xlsx'
wb = openpyxl.load_workbook(path, data_only=True)

def dump_sheet(idx):
    ws = wb[wb.sheetnames[idx]]
    lines = [f'=== SHEET {idx}: {wb.sheetnames[idx]} ===']
    for row in ws.iter_rows(values_only=True):
        cells = ['' if v is None else str(v) for v in row]
        if any(c.strip() for c in cells):
            lines.append(' | '.join(cells))
    return '\n'.join(lines)

with open(r'C:\Users\Administrator\Desktop\_xlsx_dump.txt', 'w', encoding='utf-8') as f:
    for idx in [0, 1, 3, 5, 6]:
        f.write(dump_sheet(idx))
        f.write('\n\n')

print('done')
