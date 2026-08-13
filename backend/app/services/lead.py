"""
线索池业务逻辑：录入、分配、领取、跟进、退回、转化、流失、列表过滤。
规则默认值对齐文档「待确认」：
- 保护期 15 天
- 手机号强去重；公司名软提示
- 公海领取：具备 lead:manage 或销售类角色
- 录入人可查看自己录入线索的跟进情况
"""
from __future__ import annotations

from datetime import datetime, timedelta, timezone
from typing import Optional

from fastapi import HTTPException, status
from sqlalchemy import or_
from sqlalchemy.orm import Session, joinedload
from sqlalchemy.orm.attributes import set_committed_value

from app.config import get_settings
from app.core.rbac import (
    resolve_data_scope,
    user_can,
)
from app.models.customer import Customer
from app.models.department import Department
from app.models.lead import (
    LEAD_STATUS_ASSIGNED,
    LEAD_STATUS_CONVERTED,
    LEAD_STATUS_FOLLOWING,
    LEAD_STATUS_LOST,
    LEAD_STATUS_PENDING,
    LEAD_STATUS_RETURNED,
    Lead,
    LeadFollowUp,
    LeadLog,
)
from app.models.user import User
from app.schemas.lead import (
    LeadAssignRequest,
    LeadConvertRequest,
    LeadCreate,
    LeadFollowUpCreate,
    LeadImportConfirmItem,
    LeadImportConfirmOut,
    LeadImportConfirmRequest,
    LeadImportPreviewOut,
    LeadImportPreviewRow,
    LeadImportRowIn,
    LeadLostRequest,
    LeadTransferRequest,
    LeadUpdate,
)


def _now() -> datetime:
    return datetime.now(timezone.utc)


def _as_utc(dt: Optional[datetime]) -> Optional[datetime]:
    """SQLite 读出的 DateTime 常无时区，比较前统一成 UTC aware。"""
    if dt is None:
        return None
    if dt.tzinfo is None:
        return dt.replace(tzinfo=timezone.utc)
    return dt.astimezone(timezone.utc)


def _protect_until(from_time: Optional[datetime] = None) -> datetime:
    base = _as_utc(from_time) or _now()
    days = get_settings().lead_protect_days
    return base + timedelta(days=days)


def _display_name(user: Optional[User]) -> Optional[str]:
    if not user:
        return None
    return user.real_name or user.username


def _user_name(db: Session, user_id: Optional[int]) -> Optional[str]:
    if not user_id:
        return None
    u = db.query(User).filter(User.id == user_id).first()
    if not u:
        return None
    return _display_name(u)


def _enrich_lead_logs(db: Session, logs: list[LeadLog]) -> None:
    """详情展示用真实姓名；已有日志仍可能只存了登录用户名。"""
    ids = {log.user_id for log in logs if log.user_id}
    if not ids:
        return
    names = {
        u.id: _display_name(u)
        for u in db.query(User).filter(User.id.in_(ids)).all()
    }
    for log in logs:
        name = names.get(log.user_id) if log.user_id else None
        if name and name != log.username:
            set_committed_value(log, "username", name)


def enrich_lead(db: Session, lead: Lead) -> Lead:
    """挂载展示字段（不写库）。"""
    lead.owner_name = _user_name(db, lead.owner_id)  # type: ignore[attr-defined]
    lead.creator_name = _user_name(db, lead.creator_id)  # type: ignore[attr-defined]
    protect_until = _as_utc(lead.protect_until)
    lead.is_protected = bool(  # type: ignore[attr-defined]
        protect_until
        and protect_until > _now()
        and lead.status
        in {
            LEAD_STATUS_ASSIGNED,
            LEAD_STATUS_FOLLOWING,
        }
    )
    if lead.status == LEAD_STATUS_CONVERTED and not lead.converted_opportunity_id:
        from app.services.sales_journey import resolve_converted_opportunity_id

        opp_id = resolve_converted_opportunity_id(db, lead)
        if opp_id:
            lead.converted_opportunity_id = opp_id
    return lead


def write_lead_log(
    db: Session,
    *,
    lead: Lead,
    user: Optional[User],
    action: str,
    detail: Optional[str] = None,
) -> None:
    db.add(
        LeadLog(
            lead_id=lead.id,
            user_id=user.id if user else None,
            username=_display_name(user),
            action=action,
            detail=detail,
        )
    )


def find_duplicates(
    db: Session,
    *,
    phone: Optional[str],
    company_name: Optional[str],
    credit_code: Optional[str] = None,
    company_domain: Optional[str] = None,
    exclude_id: Optional[int] = None,
) -> dict[str, list[Lead]]:
    def _base():
        q = db.query(Lead).filter(Lead.status != LEAD_STATUS_LOST)
        if exclude_id:
            q = q.filter(Lead.id != exclude_id)
        return q

    by_phone: list[Lead] = []
    by_company: list[Lead] = []
    by_credit: list[Lead] = []
    by_domain: list[Lead] = []
    if phone:
        by_phone = _base().filter(Lead.phone == phone).limit(20).all()
    if company_name:
        by_company = _base().filter(Lead.company_name == company_name).limit(20).all()
    if credit_code:
        by_credit = _base().filter(Lead.credit_code == credit_code).limit(20).all()
    if company_domain:
        by_domain = _base().filter(Lead.company_domain == company_domain).limit(20).all()
    return {
        "by_phone": by_phone,
        "by_company": by_company,
        "by_credit": by_credit,
        "by_domain": by_domain,
    }


def can_manage_lead_pool(user: User) -> bool:
    """管理层可查看/分配待分配池（对齐最终 PRD + 原型 lead-allocation）。"""
    role_codes = {r.code for r in user.roles}
    return "admin" in role_codes or user_can(user, "lead:manage")


def can_self_follow_on_create(user: User) -> bool:
    """销售录入后可直接自己跟进；其他岗位仍进管理层待分配池。"""
    role_codes = {r.code for r in (user.roles or [])}
    if "sales" in role_codes:
        return True
    # 兼容自定义角色：角色名包含"销售"或持有 lead:manage 权限视为销售
    if any("销售" in (r.name or "") for r in (user.roles or [])):
        return True
    return user_can(user, "lead:manage")


def dept_scope_ids(db: Session, user: User) -> set[int]:
    """本部门 + 所有下级部门 id（部门负责人可见整个部门树的数据）。"""
    if not user.department_id:
        return set()
    ids: set[int] = {user.department_id}
    frontier = [user.department_id]
    while frontier:
        cur = frontier.pop()
        children = db.query(Department.id).filter(Department.parent_id == cur).all()
        for (cid,) in children:
            if cid not in ids:
                ids.add(cid)
                frontier.append(cid)
    return ids


def assert_can_view(user: User, lead: Lead) -> None:
    """查看权限：待分配/已退回仅管理层；已分配按负责人/数据范围。"""
    role_codes = {r.code for r in user.roles}
    if "admin" in role_codes:
        return
    scope = resolve_data_scope(user, "lead")
    if lead.status in {LEAD_STATUS_PENDING, LEAD_STATUS_RETURNED}:
        if can_manage_lead_pool(user):
            return
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="未分配线索仅管理层可见",
        )
    if scope == "company":
        return
    if lead.owner_id == user.id or lead.creator_id == user.id:
        return
    if scope == "department" and user.department_id and lead.department_id == user.department_id:
        return
    raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="无权查看该线索")


def assert_can_operate_as_owner(user: User, lead: Lead) -> None:
    role_codes = {r.code for r in user.roles}
    if "admin" in role_codes or user_can(user, "lead:manage"):
        # 主管/管理员可操作，但仍受保护期抢单限制在 claim 里处理
        if lead.owner_id == user.id or user_can(user, "lead:manage"):
            return
    if lead.owner_id != user.id:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="仅跟进人或管理员可操作")


def create_lead(db: Session, user: User, payload: LeadCreate, *, force: bool = False) -> Lead:
    dups = find_duplicates(
        db,
        phone=payload.phone,
        company_name=payload.company_name,
        credit_code=payload.credit_code,
        company_domain=payload.company_domain,
    )
    hard = dups["by_phone"] or dups["by_credit"]
    if hard and not force:
        ids = ",".join(str(x.id) for x in (dups["by_phone"] or dups["by_credit"]))
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail=f"存在确定重复线索（ID: {ids}）。确认仍要录入请加参数 force=true",
        )

    # 销售默认录入即自跟进；非销售（或显式 self_follow=false）进待分配池
    want_self = payload.self_follow
    if want_self is None:
        want_self = can_self_follow_on_create(user)
    if want_self and not can_self_follow_on_create(user):
        raise HTTPException(status_code=403, detail="仅销售角色录入后可自己跟进，请交待分配池")

    contact_name = (payload.name or "").strip() or payload.company_name.strip()
    lead = Lead(
        name=contact_name,
        company_name=payload.company_name.strip(),
        credit_code=(payload.credit_code or "").strip() or None,
        company_domain=(payload.company_domain or "").strip() or None,
        phone=payload.phone.strip(),
        email=payload.email,
        region=payload.region,
        source=payload.source or "manual",
        source_detail=payload.source_detail,
        need_desc=payload.need_desc,
        budget=payload.budget,
        expected_deal_at=payload.expected_deal_at,
        business_type=payload.business_type,
        remark=payload.remark,
        status=LEAD_STATUS_PENDING,
        creator_id=user.id,
        department_id=user.department_id,
    )

    if want_self:
        cfg = get_settings()
        if count_protected_holds(db, user.id) >= cfg.lead_protect_hold_limit:
            raise HTTPException(
                status_code=400,
                detail=f"你保护中线索已达上限（{cfg.lead_protect_hold_limit} 条），请先清理或交待分配池",
            )
        now = _now()
        lead.owner_id = user.id
        lead.status = LEAD_STATUS_ASSIGNED
        lead.assigned_at = now
        lead.protect_until = _protect_until(now)

    db.add(lead)
    db.flush()
    detail = f"录入线索 {lead.name}"
    if want_self:
        detail += "；销售自跟进，已进入我的线索"
    else:
        detail += "；进入待分配池"
    write_lead_log(db, lead=lead, user=user, action="create", detail=detail)
    db.commit()
    db.refresh(lead)
    return enrich_lead(db, lead)


def update_lead(db: Session, user: User, lead_id: int, payload: LeadUpdate) -> Lead:
    lead = db.query(Lead).filter(Lead.id == lead_id).first()
    if not lead:
        raise HTTPException(status_code=404, detail="线索不存在")
    assert_can_view(user, lead)
    if lead.status == LEAD_STATUS_CONVERTED:
        raise HTTPException(status_code=400, detail="已转化线索不可编辑")

    changes = []
    data = payload.model_dump(exclude_unset=True)
    if "phone" in data and data["phone"]:
        dups = find_duplicates(db, phone=data["phone"], company_name=None, exclude_id=lead.id)
        if dups["by_phone"]:
            raise HTTPException(status_code=409, detail="手机号与其他线索冲突")

    for k, v in data.items():
        old = getattr(lead, k)
        if old != v:
            changes.append(f"{k}: {old} -> {v}")
            setattr(lead, k, v)

    if changes:
        write_lead_log(db, lead=lead, user=user, action="edit", detail="; ".join(changes)[:2000])
    db.commit()
    db.refresh(lead)
    return enrich_lead(db, lead)


def list_leads(
    db: Session,
    user: User,
    *,
    status: Optional[str] = None,
    keyword: Optional[str] = None,
    pool: Optional[str] = None,
    business_type: Optional[str] = None,
    page: int = 1,
    page_size: int = 20,
) -> tuple[int, list[Lead]]:
    """
    pool:
      - mine: 我负责的
      - created: 我录入的（已分配后仍可回看；未分配仅管理层可见）
      - public: 管理层线索总览（全状态：待分配/已分配/跟进/退回/转化/流失）
      - all: 按数据范围（非管理层不含未分配）
    """
    q = db.query(Lead)
    role_codes = {r.code for r in user.roles}
    is_admin = "admin" in role_codes
    is_manager = can_manage_lead_pool(user)
    scope = resolve_data_scope(user, "lead")

    if pool == "mine":
        q = q.filter(Lead.owner_id == user.id).filter(
            Lead.status.notin_([LEAD_STATUS_PENDING, LEAD_STATUS_RETURNED])
        )
    elif pool == "created":
        q = q.filter(Lead.creator_id == user.id)
        if not is_manager:
            # 录入人回看：不含仍在待分配池的线索（对齐 PRD 未分配不可见）
            q = q.filter(Lead.status.notin_([LEAD_STATUS_PENDING, LEAD_STATUS_RETURNED]))
    elif pool == "public":
        if not is_manager:
            return 0, []
        # 管理层线索总览：全状态（与「全部线索」卡片口径一致，含转化/流失）
        q = q.filter(
            Lead.status.in_(
                [
                    LEAD_STATUS_PENDING,
                    LEAD_STATUS_RETURNED,
                    LEAD_STATUS_ASSIGNED,
                    LEAD_STATUS_FOLLOWING,
                    LEAD_STATUS_CONVERTED,
                    LEAD_STATUS_LOST,
                ]
            )
        )
        # 按数据范围过滤：公司→全部；部门→本部门+自己的；个人→自己的+待分配池（供领取）
        if not is_admin and scope == "department" and user.department_id:
            q = q.filter(
                or_(
                    Lead.department_id.in_(dept_scope_ids(db, user)),
                    Lead.owner_id == user.id,
                    Lead.creator_id == user.id,
                )
            )
        elif not is_admin and scope == "personal":
            q = q.filter(
                or_(
                    Lead.owner_id == user.id,
                    Lead.creator_id == user.id,
                    Lead.status.in_([LEAD_STATUS_PENDING, LEAD_STATUS_RETURNED]),
                )
            )
    else:
        # all：按数据范围；非管理层排除未分配
        if not is_admin and scope == "personal":
            q = q.filter(or_(Lead.owner_id == user.id, Lead.creator_id == user.id))
        elif not is_admin and scope == "department" and user.department_id:
            q = q.filter(
                or_(
                    Lead.department_id.in_(dept_scope_ids(db, user)),
                    Lead.owner_id == user.id,
                    Lead.creator_id == user.id,
                )
            )
        if not is_manager:
            q = q.filter(Lead.status.notin_([LEAD_STATUS_PENDING, LEAD_STATUS_RETURNED]))

    if status:
        # 线索总览卡片快捷筛选别名
        status_groups = {
            "unassigned": [LEAD_STATUS_PENDING, LEAD_STATUS_RETURNED],
            "owned": [LEAD_STATUS_ASSIGNED, LEAD_STATUS_FOLLOWING],
        }
        if status in status_groups:
            q = q.filter(Lead.status.in_(status_groups[status]))
        elif "," in status:
            parts = [s.strip() for s in status.split(",") if s.strip()]
            q = q.filter(Lead.status.in_(parts)) if parts else q
        else:
            q = q.filter(Lead.status == status)
    if business_type:
        q = q.filter(Lead.business_type == business_type)
    if keyword:
        like = f"%{keyword}%"
        q = q.filter(
            or_(
                Lead.name.ilike(like),
                Lead.company_name.ilike(like),
                Lead.phone.ilike(like),
                Lead.need_desc.ilike(like),
                Lead.region.ilike(like),
            )
        )

    total = q.count()
    items = (
        q.order_by(Lead.updated_at.desc())
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    return total, [enrich_lead(db, x) for x in items]


def get_lead_detail(db: Session, user: User, lead_id: int) -> Lead:
    lead = (
        db.query(Lead)
        .options(joinedload(Lead.follow_ups), joinedload(Lead.logs))
        .filter(Lead.id == lead_id)
        .first()
    )
    if not lead:
        raise HTTPException(status_code=404, detail="线索不存在")
    assert_can_view(user, lead)
    # 时间倒序展示
    lead.follow_ups = sorted(lead.follow_ups or [], key=lambda x: x.id, reverse=True)
    lead.logs = sorted(lead.logs or [], key=lambda x: x.id, reverse=True)
    _enrich_lead_logs(db, lead.logs)
    return enrich_lead(db, lead)


def assign_lead(db: Session, user: User, lead_id: int, payload: LeadAssignRequest) -> Lead:
    if not can_manage_lead_pool(user):
        raise HTTPException(status_code=403, detail="无分配权限")

    lead = db.query(Lead).filter(Lead.id == lead_id).first()
    if not lead:
        raise HTTPException(status_code=404, detail="线索不存在")
    if lead.status not in {LEAD_STATUS_PENDING, LEAD_STATUS_RETURNED}:
        raise HTTPException(status_code=400, detail="仅待分配线索可分配")

    owner = db.query(User).filter(User.id == payload.owner_id, User.is_active.is_(True)).first()
    if not owner:
        raise HTTPException(status_code=400, detail="接收人不存在或已禁用")

    # 接收人保护期持有上限（对齐 LEAD-010 / LEAD-012）
    cfg = get_settings()
    if count_protected_holds(db, owner.id) >= cfg.lead_protect_hold_limit:
        raise HTTPException(
            status_code=400,
            detail=f"接收人保护中线索已达上限（{cfg.lead_protect_hold_limit} 条）",
        )

    now = _now()
    lead.owner_id = owner.id
    lead.department_id = owner.department_id or lead.department_id
    lead.status = LEAD_STATUS_ASSIGNED
    lead.assigned_at = now
    lead.protect_until = _protect_until(now)
    write_lead_log(
        db,
        lead=lead,
        user=user,
        action="assign",
        detail=f"分配给 {owner.real_name or owner.username}"
        + (f"；{payload.remark}" if payload.remark else ""),
    )
    db.commit()
    db.refresh(lead)
    return enrich_lead(db, lead)


def batch_assign_leads(
    db: Session,
    user: User,
    *,
    lead_ids: list[int],
    owner_ids: list[int],
    method: str = "average",
    assignments: Optional[list[dict]] = None,
    reason: Optional[str] = None,
) -> dict:
    """
    管理层批量分配（对齐 PRD FR-014/015 与原型 lead-allocation）。
    method=average：在 owner_ids 间轮询；manual：使用 assignments[{lead_id, owner_id}]。
    单条失败不回滚整批，返回成功/失败明细。
    """
    if not can_manage_lead_pool(user):
        raise HTTPException(status_code=403, detail="无分配权限")
    if not lead_ids:
        raise HTTPException(status_code=400, detail="请选择待分配线索")

    method = (method or "average").strip().lower()
    if method not in {"average", "manual"}:
        raise HTTPException(status_code=400, detail="分配方式仅支持 average / manual")

    owners: list[User] = []
    owner_map: dict[int, User] = {}
    if method == "average":
        if not owner_ids:
            raise HTTPException(status_code=400, detail="请至少选择一名分配人")
        for oid in owner_ids:
            u = db.query(User).filter(User.id == oid, User.is_active.is_(True)).first()
            if not u:
                raise HTTPException(status_code=400, detail=f"接收人无效或不存在：{oid}")
            owners.append(u)
            owner_map[u.id] = u
    else:
        if not assignments:
            raise HTTPException(status_code=400, detail="逐条指定时请提供 assignments")
        for row in assignments:
            oid = int(row.get("owner_id"))
            if oid not in owner_map:
                u = db.query(User).filter(User.id == oid, User.is_active.is_(True)).first()
                if not u:
                    raise HTTPException(status_code=400, detail=f"接收人无效或不存在：{oid}")
                owner_map[oid] = u

    cfg = get_settings()
    planned: dict[int, int] = {}  # lead_id -> owner_id
    if method == "average":
        for i, lid in enumerate(lead_ids):
            planned[lid] = owners[i % len(owners)].id
    else:
        by_lead = {int(x["lead_id"]): int(x["owner_id"]) for x in (assignments or [])}
        for lid in lead_ids:
            if lid not in by_lead:
                raise HTTPException(status_code=400, detail=f"线索 {lid} 未指定接收人")
            planned[lid] = by_lead[lid]

    success: list[dict] = []
    failed: list[dict] = []
    # 本批内累计持有，避免同批超限
    hold_delta: dict[int, int] = {}

    for lid, oid in planned.items():
        lead = db.query(Lead).filter(Lead.id == lid).first()
        if not lead:
            failed.append({"lead_id": lid, "reason": "线索不存在"})
            continue
        if lead.status not in {LEAD_STATUS_PENDING, LEAD_STATUS_RETURNED}:
            failed.append({"lead_id": lid, "reason": f"状态不可分配（{lead.status}）"})
            continue
        owner = owner_map[oid]
        current_hold = count_protected_holds(db, oid) + hold_delta.get(oid, 0)
        if current_hold >= cfg.lead_protect_hold_limit:
            failed.append(
                {
                    "lead_id": lid,
                    "reason": f"接收人 {owner.real_name or owner.username} 保护中已达上限",
                }
            )
            continue

        now = _now()
        lead.owner_id = owner.id
        lead.department_id = owner.department_id or lead.department_id
        lead.status = LEAD_STATUS_ASSIGNED
        lead.assigned_at = now
        lead.protect_until = _protect_until(now)
        write_lead_log(
            db,
            lead=lead,
            user=user,
            action="assign",
            detail=f"批量分配给 {owner.real_name or owner.username}"
            + (f"；方式={'逐条指定' if method == 'manual' else '平均分配'}" if method else "")
            + (f"；{reason}" if reason else ""),
        )
        hold_delta[oid] = hold_delta.get(oid, 0) + 1
        success.append(
            {
                "lead_id": lid,
                "owner_id": owner.id,
                "owner_name": owner.real_name or owner.username,
            }
        )

    db.commit()
    return {
        "success_count": len(success),
        "failed_count": len(failed),
        "success": success,
        "failed": failed,
    }


def _day_start_utc(now: Optional[datetime] = None) -> datetime:
    current = _as_utc(now) or _now()
    return current.replace(hour=0, minute=0, second=0, microsecond=0)


def count_protected_holds(db: Session, user_id: int) -> int:
    now = _now()
    return (
        db.query(Lead)
        .filter(
            Lead.owner_id == user_id,
            Lead.status.in_([LEAD_STATUS_ASSIGNED, LEAD_STATUS_FOLLOWING]),
            Lead.protect_until.isnot(None),
            Lead.protect_until > now,
        )
        .count()
    )


def count_daily_claims(db: Session, user_id: int) -> int:
    start = _day_start_utc()
    return (
        db.query(LeadLog)
        .filter(
            LeadLog.user_id == user_id,
            LeadLog.action == "claim",
            LeadLog.created_at >= start,
        )
        .count()
    )


def get_lead_quota(db: Session, user: User) -> dict:
    cfg = get_settings()
    daily_claimed = count_daily_claims(db, user.id)
    protected_count = count_protected_holds(db, user.id)
    block_reason: Optional[str] = None
    if daily_claimed >= cfg.lead_daily_claim_limit:
        block_reason = f"今日抢领已达上限（{cfg.lead_daily_claim_limit} 条）"
    elif protected_count >= cfg.lead_protect_hold_limit:
        block_reason = f"保护中线索已达上限（{cfg.lead_protect_hold_limit} 条）"
    return {
        "daily_claimed": daily_claimed,
        "daily_limit": cfg.lead_daily_claim_limit,
        "protected_count": protected_count,
        "protect_limit": cfg.lead_protect_hold_limit,
        "protect_days": cfg.lead_protect_days,
        "cooldown_hours": cfg.lead_return_cooldown_hours,
        "can_claim": block_reason is None,
        "block_reason": block_reason,
    }


def _assert_claim_quota(db: Session, user: User) -> None:
    quota = get_lead_quota(db, user)
    if not quota["can_claim"]:
        raise HTTPException(status_code=400, detail=quota["block_reason"] or "已达抢领额度上限")


def claim_lead(db: Session, user: User, lead_id: int) -> Lead:
    """公海领取。"""
    can_claim = user_can(user, "lead:manage") or any(
        r.code in {"sales", "middle_manager", "admin", "executive"} for r in user.roles
    )
    if not can_claim:
        raise HTTPException(status_code=403, detail="当前角色不可领取公海线索")

    _assert_claim_quota(db, user)

    lead = db.query(Lead).filter(Lead.id == lead_id).first()
    if not lead:
        raise HTTPException(status_code=404, detail="线索不存在")
    if lead.status not in {LEAD_STATUS_PENDING, LEAD_STATUS_RETURNED}:
        raise HTTPException(status_code=400, detail="仅公海线索可领取")

    # 冷静期：刚退回的短时间内不可被原跟进人以外的人反复抢（简化：退回后仍在 protect_until 内且非原主不可领）
    protect_until = _as_utc(lead.protect_until)
    if (
        lead.status == LEAD_STATUS_RETURNED
        and protect_until
        and protect_until > _now()
        and lead.owner_id
        and lead.owner_id != user.id
    ):
        raise HTTPException(status_code=400, detail="退回冷静期内，暂不可被他人领取")

    now = _now()
    lead.owner_id = user.id
    lead.department_id = user.department_id or lead.department_id
    lead.status = LEAD_STATUS_ASSIGNED
    lead.assigned_at = now
    lead.protect_until = _protect_until(now)
    write_lead_log(db, lead=lead, user=user, action="claim", detail="从公海领取")
    db.commit()
    db.refresh(lead)
    return enrich_lead(db, lead)


def return_to_pool(
    db: Session,
    user: User,
    lead_id: int,
    reason: Optional[str] = None,
    *,
    reason_type: Optional[str] = None,
) -> Lead:
    lead = db.query(Lead).filter(Lead.id == lead_id).first()
    if not lead:
        raise HTTPException(status_code=404, detail="线索不存在")
    assert_can_operate_as_owner(user, lead)
    if lead.status in {LEAD_STATUS_CONVERTED, LEAD_STATUS_LOST}:
        raise HTTPException(status_code=400, detail="当前状态不可退回")

    now = _now()
    lead.status = LEAD_STATUS_RETURNED
    lead.protect_until = now + timedelta(hours=get_settings().lead_return_cooldown_hours)
    detail_parts = []
    if reason_type:
        detail_parts.append(f"类型:{reason_type}")
    if reason:
        detail_parts.append(reason)
    write_lead_log(
        db,
        lead=lead,
        user=user,
        action="return",
        detail="；".join(detail_parts) if detail_parts else "退回公海",
    )
    # owner_id 保留为原跟进人，便于冷静期判断；列表公海仍按 status 筛
    db.commit()
    db.refresh(lead)
    return enrich_lead(db, lead)


def transfer_lead(db: Session, user: User, lead_id: int, payload: LeadTransferRequest) -> Lead:
    lead = db.query(Lead).filter(Lead.id == lead_id).first()
    if not lead:
        raise HTTPException(status_code=404, detail="线索不存在")
    assert_can_operate_as_owner(user, lead)
    if lead.status in {LEAD_STATUS_CONVERTED, LEAD_STATUS_LOST, LEAD_STATUS_PENDING}:
        raise HTTPException(status_code=400, detail="当前状态不可流转")

    owner = db.query(User).filter(User.id == payload.owner_id, User.is_active.is_(True)).first()
    if not owner:
        raise HTTPException(status_code=400, detail="转入人不存在或已禁用")

    now = _now()
    old = _user_name(db, lead.owner_id)
    lead.owner_id = owner.id
    lead.department_id = owner.department_id or lead.department_id
    lead.status = LEAD_STATUS_ASSIGNED
    lead.assigned_at = now
    lead.protect_until = _protect_until(now)
    write_lead_log(
        db,
        lead=lead,
        user=user,
        action="transfer",
        detail=f"{old} → {owner.real_name or owner.username}"
        + (f"；{payload.reason}" if payload.reason else ""),
    )
    db.commit()
    db.refresh(lead)
    return enrich_lead(db, lead)


def add_follow_up(db: Session, user: User, lead_id: int, payload: LeadFollowUpCreate) -> LeadFollowUp:
    lead = db.query(Lead).filter(Lead.id == lead_id).first()
    if not lead:
        raise HTTPException(status_code=404, detail="线索不存在")
    assert_can_operate_as_owner(user, lead)
    if lead.status in {LEAD_STATUS_CONVERTED, LEAD_STATUS_LOST, LEAD_STATUS_PENDING, LEAD_STATUS_RETURNED}:
        raise HTTPException(status_code=400, detail="当前状态不可跟进")

    follow_at = payload.follow_at or _now()
    fu = LeadFollowUp(
        lead_id=lead.id,
        user_id=user.id,
        follow_at=follow_at,
        method=payload.method,
        content=payload.content,
        customer_feedback=payload.customer_feedback,
        result=payload.result,
        next_follow_at=payload.next_follow_at,
    )
    db.add(fu)
    lead.last_followed_at = follow_at
    lead.status = LEAD_STATUS_FOLLOWING

    if payload.result == "return":
        lead.status = LEAD_STATUS_RETURNED
        lead.protect_until = _now() + timedelta(hours=get_settings().lead_return_cooldown_hours)
    elif payload.result == "lost":
        lead.status = LEAD_STATUS_LOST
        lead.lost_reason = payload.customer_feedback or payload.content
        lead.lost_at = follow_at

    method_label = {
        "phone": "电话",
        "wechat": "微信",
        "email": "邮件",
        "meeting": "面谈",
        "conference": "会议",
    }.get(payload.method, payload.method)
    result_label = {
        "advance": "推进",
        "keep": "保持",
        "return": "退回",
        "lost": "流失",
    }.get(payload.result, payload.result)
    write_lead_log(
        db,
        lead=lead,
        user=user,
        action="follow",
        detail=f"{method_label}/{result_label}：{payload.content[:200]}",
    )
    db.commit()
    db.refresh(fu)
    return fu


def convert_lead(db: Session, user: User, lead_id: int, payload: LeadConvertRequest) -> dict:
    from app.models.opportunity import OPP_STAGE_NEED, OPP_STAGES, Opportunity, OpportunityActivity
    from app.models.opportunity import OPP_STAGE_LABEL
    from app.services.opportunity import _gen_opportunity_no
    from app.services import platform as platform_service

    lead = db.query(Lead).filter(Lead.id == lead_id).first()
    if not lead:
        raise HTTPException(status_code=404, detail="线索不存在")
    assert_can_operate_as_owner(user, lead)
    if lead.status in {LEAD_STATUS_CONVERTED, LEAD_STATUS_LOST}:
        raise HTTPException(status_code=400, detail="当前状态不可转化")

    business_type = payload.business_type or lead.business_type or "other"
    if business_type not in platform_service.business_type_values(db, enabled_only=False):
        business_type = "other"
    stage = payload.opportunity_stage or OPP_STAGE_NEED
    if stage not in OPP_STAGES:
        raise HTTPException(status_code=400, detail="无效的商机阶段")

    customer_name = payload.customer_name or lead.company_name or lead.name
    customer = Customer(
        name=customer_name,
        contact_name=lead.name,
        phone=lead.phone,
        email=lead.email,
        source=lead.source or "lead_convert",
        status="potential",
        owner_id=lead.owner_id or user.id,
        creator_id=user.id,
        department_id=lead.department_id or user.department_id,
        source_lead_id=lead.id,
        remark=payload.remark or f"由线索#{lead.id}转化",
    )
    db.add(customer)
    db.flush()

    opp_title = (payload.opportunity_title or customer_name).strip()
    opp = Opportunity(
        opportunity_no=_gen_opportunity_no(db),
        title=opp_title,
        customer_id=customer.id,
        source_lead_id=lead.id,
        business_type=business_type,
        stage=stage,
        expected_amount=payload.expected_amount if payload.expected_amount is not None else (lead.budget or 0),
        currency="CNY",
        owner_id=lead.owner_id or user.id,
        creator_id=user.id,
        department_id=lead.department_id or user.department_id,
        requirement_summary=payload.requirement_summary or lead.need_desc,
        remark=f"由线索#{lead.id}转化",
    )
    db.add(opp)
    db.flush()
    db.add(
        OpportunityActivity(
            opportunity_id=opp.id,
            user_id=user.id,
            activity_type="create",
            content=f"线索转化创建商机，阶段：{OPP_STAGE_LABEL.get(stage, stage)}",
            to_stage=stage,
        )
    )

    now = _now()
    lead.status = LEAD_STATUS_CONVERTED
    lead.converted_customer_id = customer.id
    lead.converted_opportunity_id = opp.id
    lead.converted_at = now
    write_lead_log(
        db,
        lead=lead,
        user=user,
        action="convert",
        detail=f"转化为客户#{customer.id}、商机#{opp.id} {opp.title}",
    )
    db.commit()
    db.refresh(lead)
    return {
        "lead": enrich_lead(db, lead),
        "customer_id": customer.id,
        "opportunity_id": opp.id,
    }


def mark_lost(db: Session, user: User, lead_id: int, payload: LeadLostRequest) -> Lead:
    lead = db.query(Lead).filter(Lead.id == lead_id).first()
    if not lead:
        raise HTTPException(status_code=404, detail="线索不存在")
    assert_can_operate_as_owner(user, lead)
    if lead.status == LEAD_STATUS_CONVERTED:
        raise HTTPException(status_code=400, detail="已转化线索不可标记流失")

    now = _now()
    lead.status = LEAD_STATUS_LOST
    lead.lost_reason = payload.reason
    lead.lost_at = now
    write_lead_log(db, lead=lead, user=user, action="lost", detail=payload.reason)
    db.commit()
    db.refresh(lead)
    return enrich_lead(db, lead)


def lead_stats(db: Session, user: User) -> dict:
    from datetime import timedelta

    def _count(status: Optional[str] = None, pool: Optional[str] = None) -> int:
        total, _ = list_leads(db, user, status=status, pool=pool, page=1, page_size=1)
        return total

    now = _now()
    day_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
    month_start = day_start.replace(day=1)
    protect_soon = now + timedelta(days=3)

    today_created = db.query(Lead).filter(Lead.created_at >= day_start).count()
    today_assigned = (
        db.query(Lead)
        .filter(Lead.assigned_at.isnot(None), Lead.assigned_at >= day_start)
        .count()
    )
    following_mine = (
        db.query(Lead)
        .filter(Lead.owner_id == user.id, Lead.status == LEAD_STATUS_FOLLOWING)
        .count()
    )
    protect_expiring = (
        db.query(Lead)
        .filter(
            Lead.owner_id == user.id,
            Lead.protect_until.isnot(None),
            Lead.protect_until >= now,
            Lead.protect_until <= protect_soon,
        )
        .count()
    )
    converted_month = (
        db.query(Lead)
        .filter(
            Lead.owner_id == user.id,
            Lead.status == LEAD_STATUS_CONVERTED,
            Lead.converted_at.isnot(None),
            Lead.converted_at >= month_start,
        )
        .count()
    )

    pending_pool = 0
    if can_manage_lead_pool(user):
        _scope = resolve_data_scope(user, "lead")
        _is_admin = "admin" in {r.code for r in user.roles}
        pq = db.query(Lead).filter(Lead.status.in_([LEAD_STATUS_PENDING, LEAD_STATUS_RETURNED]))
        # 与列表口径一致：部门→本部门待分配；公司/个人(销售可领公海)→全部
        if not _is_admin and _scope == "department" and user.department_id:
            pq = pq.filter(
                or_(
                    Lead.department_id.in_(dept_scope_ids(db, user)),
                    Lead.owner_id == user.id,
                    Lead.creator_id == user.id,
                )
            )
        pending_pool = pq.count()

    return {
        "total": _count(pool="all"),
        "pending_assign": _count(status=LEAD_STATUS_PENDING, pool="all"),
        "assigned": _count(status=LEAD_STATUS_ASSIGNED, pool="all"),
        "following": _count(status=LEAD_STATUS_FOLLOWING, pool="all"),
        "converted": _count(status=LEAD_STATUS_CONVERTED, pool="all"),
        "returned": _count(status=LEAD_STATUS_RETURNED, pool="all"),
        "lost": _count(status=LEAD_STATUS_LOST, pool="all"),
        "public_pool": pending_pool,
        "today_created": today_created,
        "today_assigned": today_assigned,
        "following_mine": following_mine,
        "protect_expiring": protect_expiring,
        "converted_month": converted_month,
    }


# ---- 批量导入（Excel .xlsx / CSV） ----

IMPORT_MAX_ROWS = 200
IMPORT_HEADERS = [
    "客户主体",
    "联系电话",
    "联系人",
    "统一社会信用代码",
    "企业域名",
    "需求方向",
    "需求说明",
    "备注",
]
_IMPORT_SAMPLE_ROW = [
    "示例科技有限公司",
    "13800138000",
    "张三",
    "",
    "example.com",
    "AI产品销售",
    "需要智能客服方案",
    "",
]
_HEADER_ALIASES = {
    "客户主体": "company_name",
    "公司名称": "company_name",
    "公司名": "company_name",
    "联系电话": "phone",
    "手机号": "phone",
    "电话": "phone",
    "联系人": "name",
    "姓名": "name",
    "统一社会信用代码": "credit_code",
    "信用代码": "credit_code",
    "企业域名": "company_domain",
    "域名": "company_domain",
    "需求方向": "business_type",
    "业务类型": "business_type",
    "需求说明": "need_desc",
    "备注": "remark",
}


def build_import_template_csv() -> bytes:
    """UTF-8 BOM CSV，Excel 打开中文不乱码。"""
    import csv
    import io

    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(IMPORT_HEADERS)
    writer.writerow(_IMPORT_SAMPLE_ROW)
    return ("\ufeff" + buf.getvalue()).encode("utf-8")


def build_import_template_xlsx() -> bytes:
    """标准 Excel 模板（.xlsx）。"""
    import io

    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "线索导入"
    ws.append(IMPORT_HEADERS)
    ws.append(_IMPORT_SAMPLE_ROW)
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 18
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _normalize_header(cell: str) -> str:
    return (cell or "").strip().lstrip("\ufeff")


def _cell_str(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _resolve_business_type(db: Session, raw: str) -> tuple[Optional[str], str]:
    from app.services import platform as platform_service

    text = (raw or "").strip() or "ai_product"
    items = platform_service.list_business_type_items(db, enabled_only=False)
    by_value = {x["value"]: x["label"] for x in items}
    by_label = {x["label"]: x["value"] for x in items}
    if text in by_value:
        return text, by_value[text]
    if text in by_label:
        code = by_label[text]
        return code, text
    enabled = platform_service.business_type_values(db, enabled_only=True)
    if text in enabled:
        return text, by_value.get(text, text)
    return None, text


def _rows_from_matrix(matrix: list[list[str]]) -> list[dict]:
    if not matrix:
        raise HTTPException(status_code=400, detail="文件无内容")
    headers = [_normalize_header(h) for h in matrix[0]]
    field_map: dict[int, str] = {}
    for idx, h in enumerate(headers):
        key = _HEADER_ALIASES.get(h)
        if key:
            field_map[idx] = key
    if "company_name" not in field_map.values() or "phone" not in field_map.values():
        raise HTTPException(
            status_code=400,
            detail="模板缺少必填列：客户主体、联系电话。请下载标准模板后填写",
        )
    parsed: list[dict] = []
    for i, raw in enumerate(matrix[1:], start=2):
        if not any((c or "").strip() for c in raw):
            continue
        item: dict = {"row_no": i}
        for idx, key in field_map.items():
            val = raw[idx].strip() if idx < len(raw) else ""
            item[key] = val or None
        parsed.append(item)
    if not parsed:
        raise HTTPException(status_code=400, detail="没有可导入的数据行")
    if len(parsed) > IMPORT_MAX_ROWS:
        raise HTTPException(
            status_code=400,
            detail=f"单次最多导入 {IMPORT_MAX_ROWS} 条，当前 {len(parsed)} 条",
        )
    return parsed


def parse_import_csv(content: bytes) -> list[dict]:
    import csv
    import io

    if not content:
        raise HTTPException(status_code=400, detail="文件为空")
    text = content.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text))
    matrix = [[_cell_str(c) for c in row] for row in reader]
    return _rows_from_matrix(matrix)


def parse_import_xlsx(content: bytes) -> list[dict]:
    import io

    from openpyxl import load_workbook

    if not content:
        raise HTTPException(status_code=400, detail="文件为空")
    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=400, detail=f"无法读取 Excel 文件：{exc}") from exc
    ws = wb.active
    matrix: list[list[str]] = []
    for row in ws.iter_rows(values_only=True):
        matrix.append([_cell_str(c) for c in row])
    wb.close()
    return _rows_from_matrix(matrix)


def parse_import_file(content: bytes, filename: str = "") -> list[dict]:
    name = (filename or "").lower()
    if name.endswith(".xlsx") or name.endswith(".xlsm"):
        return parse_import_xlsx(content)
    if name.endswith(".xls"):
        raise HTTPException(
            status_code=400,
            detail="暂不支持旧版 .xls，请另存为 .xlsx 或 CSV UTF-8 后再上传",
        )
    if name.endswith(".csv") or name.endswith(".txt") or not name:
        # 无后缀时先按 CSV，失败再尝试 xlsx（部分浏览器丢扩展名）
        try:
            return parse_import_csv(content)
        except HTTPException:
            if content[:2] == b"PK":
                return parse_import_xlsx(content)
            raise
    if content[:2] == b"PK":
        return parse_import_xlsx(content)
    raise HTTPException(status_code=400, detail="请上传 Excel（.xlsx）或 CSV 文件")


def preview_lead_import(
    db: Session, content: bytes, *, filename: str = ""
) -> LeadImportPreviewOut:
    raw_rows = parse_import_file(content, filename)
    phone_seen: dict[str, int] = {}
    credit_seen: dict[str, int] = {}
    out_rows: list[LeadImportPreviewRow] = []
    ok = soft = hard = err = 0

    for item in raw_rows:
        row_no = int(item["row_no"])
        company = (item.get("company_name") or "").strip()
        phone = (item.get("phone") or "").strip()
        credit = (item.get("credit_code") or "").strip() or None
        domain = (item.get("company_domain") or "").strip() or None
        name = (item.get("name") or "").strip() or None
        need_desc = (item.get("need_desc") or "").strip() or None
        remark = (item.get("remark") or "").strip() or None
        bt_raw = item.get("business_type") or "ai_product"
        bt_code, bt_label = _resolve_business_type(db, str(bt_raw))

        status_code = "ok"
        message = "可导入"
        can_import = True
        force_required = False

        if not company or not phone:
            status_code = "error"
            message = "客户主体与联系电话为必填"
            can_import = False
            err += 1
        elif bt_code is None:
            status_code = "error"
            message = f"需求方向无效：{bt_label}"
            can_import = False
            err += 1
        else:
            if phone in phone_seen:
                status_code = "error"
                message = f"与文件内第 {phone_seen[phone]} 行手机号重复"
                can_import = False
                err += 1
            elif credit and credit in credit_seen:
                status_code = "error"
                message = f"与文件内第 {credit_seen[credit]} 行信用代码重复"
                can_import = False
                err += 1
            else:
                phone_seen[phone] = row_no
                if credit:
                    credit_seen[credit] = row_no
                dups = find_duplicates(
                    db,
                    phone=phone,
                    company_name=company,
                    credit_code=credit,
                    company_domain=domain,
                )
                if dups["by_phone"] or dups["by_credit"]:
                    status_code = "hard"
                    ids = ",".join(
                        str(x.id) for x in (dups["by_phone"] or dups["by_credit"])
                    )
                    message = f"确定重复（已有线索 ID: {ids}），勾选强制后可导入"
                    force_required = True
                    hard += 1
                elif dups["by_company"] or dups["by_domain"]:
                    status_code = "soft"
                    message = "疑似重复（公司名或域名相近），可导入并留痕"
                    soft += 1
                else:
                    ok += 1

        out_rows.append(
            LeadImportPreviewRow(
                row_no=row_no,
                company_name=company,
                phone=phone,
                name=name,
                credit_code=credit,
                company_domain=domain,
                business_type=bt_code or "ai_product",
                business_type_label=bt_label,
                need_desc=need_desc,
                remark=remark,
                status=status_code,
                message=message,
                can_import=can_import,
                force_required=force_required,
            )
        )

    return LeadImportPreviewOut(
        total=len(out_rows),
        ok_count=ok,
        soft_count=soft,
        hard_count=hard,
        error_count=err,
        rows=out_rows,
    )


def confirm_lead_import(
    db: Session,
    user: User,
    payload: LeadImportConfirmRequest,
) -> LeadImportConfirmOut:
    if len(payload.rows) > IMPORT_MAX_ROWS:
        raise HTTPException(status_code=400, detail=f"单次最多导入 {IMPORT_MAX_ROWS} 条")

    items: list[LeadImportConfirmItem] = []
    success = failed = skipped = 0
    for row in payload.rows:
        try:
            bt_code, _ = _resolve_business_type(db, row.business_type)
            if not bt_code:
                raise HTTPException(status_code=400, detail="需求方向无效")
            create_payload = LeadCreate(
                name=row.name,
                company_name=row.company_name,
                credit_code=row.credit_code,
                company_domain=row.company_domain,
                phone=row.phone,
                business_type=bt_code,
                need_desc=row.need_desc,
                remark=row.remark,
                source="batch_import",
                source_detail=f"批量导入第{row.row_no}行",
                self_follow=payload.self_follow,
            )
            lead = create_lead(db, user, create_payload, force=row.force)
            success += 1
            items.append(
                LeadImportConfirmItem(
                    row_no=row.row_no,
                    ok=True,
                    lead_id=lead.id,
                    message="已导入",
                )
            )
        except HTTPException as exc:
            failed += 1
            detail = exc.detail if isinstance(exc.detail, str) else str(exc.detail)
            items.append(
                LeadImportConfirmItem(row_no=row.row_no, ok=False, message=detail)
            )
        except Exception as exc:  # noqa: BLE001
            failed += 1
            items.append(
                LeadImportConfirmItem(row_no=row.row_no, ok=False, message=str(exc))
            )

    return LeadImportConfirmOut(
        success_count=success,
        failed_count=failed,
        skipped_count=skipped,
        items=items,
    )
